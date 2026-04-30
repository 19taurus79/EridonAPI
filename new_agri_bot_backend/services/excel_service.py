import io
from typing import List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from aiogram.types import BufferedInputFile

from ..models import DeliveryRequest
from ..config import bot, logger, SEND_NOTIFICATIONS

def export_delivery_to_excel(data: DeliveryRequest) -> Workbook:
    """
    Генерує Excel-файл для заявки на доставку.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Доставка"

    header_font = Font(bold=True, size=14)
    ws.append(["Менеджер", data.manager])
    ws["A1"].font = header_font
    ws["B1"].font = Font(bold=True)

    ws.append(["Контрагент", data.client])
    ws["A2"].font = header_font
    ws["B2"].font = Font(bold=True)

    ws.append(["Адреса", data.address])
    ws["A3"].font = header_font
    ws["B3"].font = Font(bold=True)

    ws.append(["Контакт", data.contact])
    ws["A4"].font = header_font
    ws["B4"].font = Font(bold=True)

    ws.append(["Телефон", data.phone])
    ws["A5"].font = header_font
    ws["B5"].font = Font(bold=True)

    ws.append(["Дата", data.date])
    ws["A6"].font = header_font
    ws["B6"].font = Font(bold=True)

    ws.append(["Коментар", data.comment or ""])
    ws["A7"].font = header_font
    ws["B7"].font = Font(bold=True)

    ws.append([])

    header_fill = PatternFill(start_color="DDEBF7", fill_type="solid")
    title_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.append(["Доповнення", "Товар", "Кількість"])
    row = ws.max_row
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for order in data.orders:
        for item in order.items:
            ws.append([order.order, item.product, item.quantity])
            main_row = ws.max_row
            main_bold_font = Font(bold=True)
            ws[f"A{main_row}"].font = main_bold_font
            ws[f"A{main_row}"].alignment = Alignment(horizontal="left")
            ws[f"B{main_row}"].font = main_bold_font
            ws[f"B{main_row}"].alignment = Alignment(horizontal="left")
            ws[f"C{main_row}"].font = main_bold_font
            ws[f"C{main_row}"].alignment = Alignment(horizontal="right")
            for col in range(1, 4):
                ws.cell(row=main_row, column=col).border = thin_border

            if item.parties and item.parties[0].moved_q > 0:
                for party in item.parties:
                    ws.append(["", f"  ↳ {party.party}", party.moved_q])
                    party_row = ws.max_row
                    party_font = Font(italic=True, size=11)
                    ws[f"B{party_row}"].font = party_font
                    ws[f"B{party_row}"].alignment = Alignment(horizontal="left")
                    ws[f"C{party_row}"].font = party_font
                    ws[f"C{party_row}"].alignment = Alignment(horizontal="left")
                    for col in range(1, 4):
                        ws.cell(row=party_row, column=col).border = thin_border

    last_row = ws.max_row
    for col in range(1, 4):
        ws.cell(row=last_row, column=col).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="double"),
        )

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

async def send_delivery_excel_report(data: DeliveryRequest, chat_ids: List[int]):
    """
    Генерує та надсилає Excel-звіт про доставку в Telegram.
    """
    if not SEND_NOTIFICATIONS:
        logger.info("🔇 Сповіщення вимкнено. Excel-звіт не буде надіслано.")
        return

    try:
        wb = export_delivery_to_excel(data)
        
        # Збереження в буфер
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        safe_manager = data.manager.replace(" ", "_")
        filename = f"Доставка_{safe_manager}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        input_file = BufferedInputFile(excel_buffer.getvalue(), filename=filename)
        
        for chat_id in chat_ids:
            try:
                await bot.send_document(chat_id=chat_id, document=input_file)
                logger.info(f"✅ Excel-звіт надіслано користувачу {chat_id}")
            except Exception as e:
                logger.error(f"❌ Помилка надсилання Excel користувачу {chat_id}: {e}")
                
    except Exception as e:
        logger.error(f"❌ Помилка генерації або надсилання Excel-звіту: {e}")
